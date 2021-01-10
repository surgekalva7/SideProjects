import requests
import smtplib
import os
from glob import glob
from bs4 import BeautifulSoup
import pandas as pd
from datetime import datetime
from time import sleep
from email.message import EmailMessage
from openpyxl import load_workbook


class AmazonScraper:
    #this is the constructor for the class
    def __init__(self):
        #header is needed for all amazon searches
        self.HEADERS = ({'User-Agent':
                'Mozilla/5.0 (Windows NT 6.1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/41.0.2228.0 Safari/537.36',
                'Accept-Language': 'en-US, en;q=0.5'})
    
    #this is the method that will send an email
    def email_alert(self,subject,body,to):
        msg = EmailMessage()
        msg.set_content(body)
        msg['subject'] = subject
        msg['to']  = to

        user = "surgekalva7@gmail.com"
        msg['from'] = user    
        password = "gozdouruqihejpel"

        server = smtplib.SMTP("smtp.gmail.com",587)
        server.starttls()
        server.login(user,password)
        server.send_message(msg)
        server.quit()

    def searchProductList(self,email):
        #This is the location of the urls.csv file
        prod = pd.read_csv('/Users/srujankalva/Desktop/Jobs:Internships/PythonProjects/SaveMoney/urls.csv')
        #creates a map of the url and the prices
        urls_fromcsv = prod.url
        urls = urls_fromcsv.copy()

        prices_fromcsv = prod.price
        prices = prices_fromcsv.copy()

        initial_length = urls.size

        row = 1
        #this is the location of the data being stored
        workbook = load_workbook('/Users/srujankalva/Desktop/Jobs:Internships/PythonProjects/SaveMoney/SearchHistory.xlsx')
        #finds the sheet of the excel page that is being used
        sheet = workbook.active
        #header of the excel page
        sheet.cell(row,2, value = 'date')
        sheet.cell(row,3, value = 'url')
        sheet.cell(row,4, value = 'product name')
        sheet.cell(row,5, value = 'current price')
        sheet.cell(row,6, value = 'ask price')
        sheet.cell(row,7, value = 'stock')
        #runs until all the urls are found to be below the ask price
        while(urls.size>0):
            count = 0
            now = datetime.now().strftime('%Y-%m-%d %Hh%Mm')
            current_length = urls.size
            runs = 0
            #runs for all urls that have not yet been found to be below the ask price
            while(count<initial_length and runs<current_length):
                #checks to make sure that the url is not empty        
                if not urls[count]=='':
                    runs+=1                   
                    page = requests.get(urls[count],headers=self.HEADERS)

                    #turns this into understandable code
                    soup = BeautifulSoup(page.content, features="lxml")

                    #attributes of the product found from the html page
                    product_title = soup.find(id='productTitle').get_text().strip()
                    try:
                        product_price = float(soup.find(id='priceblock_ourprice').get_text().replace('$', '').replace(',', '').strip())
                    except:
                        product_price = ''

                    try:
                        soup.select('#availability .a-color-state')[0].get_text().strip()
                        product_stock = 'Out of Stock'
                    except:
                        product_stock = 'Available'
                    #what is going to be logged
                    product_url = urls[count]
                    product_ask_price = prices[count]
                    #should determine if there should be an email sent or not
                    try:
                        if(product_stock=='Available' and float(product_price)<float(prices[count])):
                            self.email_alert(product_title, urls[count],email)
                            urls[count] = ''
                            prices[count] = ''
                            
                    except:
                        pass  
                    #inputs the data into the excel file                    
                    row +=1
                    sheet.cell(row,1,value=row-1)
                    sheet.cell(row,2,value=now)
                    sheet.cell(row,3,value=product_url)
                    sheet.cell(row,4,value=product_title)
                    sheet.cell(row,5,value=product_price)
                    sheet.cell(row,6,value=product_ask_price)
                    sheet.cell(row,7,value=product_stock)
                    workbook.save('/Users/srujankalva/Desktop/Jobs:Internships/PythonProjects/SaveMoney/SearchHistory.xlsx')
                #moves onto the next url
                count+=1
            #waits to run the while loop again(1 hour)
            sleep(3600)
        print('end of search')
